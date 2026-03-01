"""
excel_evidence_gen.py

output/screenshots/ に保存されたスクリーンショット PNG を読み込み、
1ファイル1シートの Excel エビデンスファイルを生成するスクリプト。

設計意図:
    Playwright テストが生成したスクリーンショットをエビデンスとして提出する際、
    1枚ずつ Excel に貼り付ける手作業を排除する。
    シート順は PNG ファイル名の昇順とし、TC-001, TC-002, ... の自然順に一致させる。

使用ライブラリ:
    - openpyxl: Excel ファイルの生成・シート操作・画像挿入
    - Pillow:   元画像のピクセルサイズ取得（openpyxl の PNG 読み込みにも内部利用される）

実行例:
    # デフォルトパスで実行（プロジェクトルートから）
    python scripts/excel_evidence_gen.py

    # パスを明示する場合
    python scripts/excel_evidence_gen.py \\
        --screenshots-dir ./output/screenshots \\
        --output ./output/evidence.xlsx
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font
from PIL import Image as PilImage


# Excel のシート名上限（仕様）
SHEET_NAME_MAX_LEN = 31

# 画像の最大幅（px）。フルページ SS は幅 1280px 前後のため、そのまま挿入すると
# Excel 上で横に大きすぎる。800px を上限としてアスペクト比を維持してリサイズする。
IMAGE_MAX_WIDTH_PX = 800

# シート名マッピング。
# キー: PNG ファイルの stem（拡張子なしファイル名）
# 値:   Excel に表示するシート名
# マッピングにないファイルはファイル名の stem をそのままシート名として使用する。
SHEET_NAME_MAPPING: dict[str, str] = {
    "TC-001_直接遷移_条件なし検索": "01_直接アクセス確認",
    "TC-002_メニュー経由_セッション非依存画面_条件なし検索": "02_メニュー経由確認",
    "TC-003_メニュー経由_セッション依存画面_条件なし検索": "03_引継検索_メニュー経由確認",
    "TC-004_引継ぎ条件クリア後に全件検索": "04_引継条件クリア_全件検索確認",
    "TC-005_フリーワードのみ検索": "05_フリーワード指定",
    "TC-006_部門コードのみ検索": "06_部門コード指定",
    "TC-007_複数条件で検索": "07_複合条件指定",
}


def parse_args() -> argparse.Namespace:
    """CLI 引数をパースして返す。

    デフォルト値はスクリプトファイルの位置（scripts/）を起点に
    プロジェクトルートを算出した絶対パスを使用する。
    これにより、カレントディレクトリに依存せずどこからでも実行できる。
    """
    # このスクリプトは scripts/ に配置されているため、親ディレクトリがプロジェクトルート
    project_root = Path(__file__).resolve().parent.parent
    default_screenshots_dir = project_root / "output" / "screenshots"
    default_output = project_root / "output" / "evidence.xlsx"

    parser = argparse.ArgumentParser(
        description="Playwright スクリーンショットを Excel エビデンスに変換する"
    )
    parser.add_argument(
        "--screenshots-dir",
        default=str(default_screenshots_dir),
        help=f"PNG を格納したディレクトリ（デフォルト: {default_screenshots_dir}）",
    )
    parser.add_argument(
        "--output",
        default=str(default_output),
        help=f"出力 Excel ファイルのパス（デフォルト: {default_output}）",
    )
    return parser.parse_args()


def collect_png_files(screenshots_dir: Path) -> list[Path]:
    """指定ディレクトリから PNG ファイルをファイル名昇順で収集する。

    ファイル名昇順は TC-001, TC-002, ... のテストケース番号順に自然に一致する。
    ディレクトリが存在しない、または PNG が 0 件の場合は早期終了させるため
    呼び出し元でエラーチェックを行う。
    """
    if not screenshots_dir.is_dir():
        print(
            f"エラー: ディレクトリが見つかりません: {screenshots_dir}", file=sys.stderr
        )
        sys.exit(1)

    png_files = sorted(screenshots_dir.glob("*.png"), key=lambda p: p.name)

    if not png_files:
        print(
            f"エラー: PNG ファイルが見つかりません: {screenshots_dir}", file=sys.stderr
        )
        sys.exit(1)

    return png_files


def calc_display_size(png_path: Path) -> tuple[int, int]:
    """Pillow で元画像のサイズを取得し、最大幅 IMAGE_MAX_WIDTH_PX を上限に
    アスペクト比を維持した表示サイズ（width, height）を返す。

    openpyxl の Image.width / Image.height はピクセル単位で指定できる。
    元画像が上限幅以下の場合はそのままのサイズを返す。
    """
    with PilImage.open(png_path) as img:
        orig_w, orig_h = img.size

    if orig_w <= IMAGE_MAX_WIDTH_PX:
        return orig_w, orig_h

    # 縮小比率を幅基準で計算し、高さも同率で縮小する
    scale = IMAGE_MAX_WIDTH_PX / orig_w
    return IMAGE_MAX_WIDTH_PX, int(orig_h * scale)


def make_sheet_name(stem: str, name_map: dict[str, str]) -> str:
    """ファイルの stem（拡張子なしファイル名）を Excel シート名に変換する。

    name_map に stem が存在する場合はその値をシート名として使用する。
    存在しない場合は stem をそのまま使用する（フォールバック）。
    どちらのケースでも Excel のシート名上限（31 文字）を超える場合は末尾を切り捨てる。
    シート名が切り捨てられた場合でも A1 セルにフルネームを記載しているため
    元のファイル名は参照可能。
    """
    display_name = name_map.get(stem, stem)  # マッピングにあれば置換、なければ stem
    return display_name[:SHEET_NAME_MAX_LEN]


def add_screenshot_sheet(
    wb: Workbook, png_path: Path, name_map: dict[str, str]
) -> None:
    """1 枚の PNG に対応する Excel シートを Workbook に追加する。

    name_map を参照して stem → シート名の変換を行う。マッピングにない stem は
    そのままシート名として使用する。A1 には元のファイル名を記載するため、
    シート名が別名に変わっても元ファイルとの対応を確認できる。

    レイアウト:
        A1: ファイル名（太字）— シート名が切り詰められた場合や別名の場合の参照用
        A2: スクリーンショット画像（アンカー A2）

    行・列のサイズを概算で調整し、視認性を向上させる。
    """
    sheet_name = make_sheet_name(png_path.stem, name_map)
    ws = wb.create_sheet(title=sheet_name)

    # A1: ファイル名を太字で記入する（シート名切り捨て時の照合用）
    ws["A1"] = png_path.name
    ws["A1"].font = Font(bold=True)

    # 画像の表示サイズを計算し、openpyxl の Image オブジェクトに反映する
    display_w, display_h = calc_display_size(png_path)
    img = XlImage(str(png_path))
    img.width = display_w
    img.height = display_h

    # A2 をアンカーとして画像を挿入する。
    # セルサイズは変更せず、画像はセルをはみ出して配置される（Excel の通常の画像挿入と同じ挙動）。
    ws.add_image(img, "A2")

    ws.row_dimensions[1].height = 18  # ヘッダー行（ファイル名テキスト）のみ調整


def main() -> None:
    """メインエントリポイント。CLI 引数に基づいて Excel エビデンスを生成する。

    処理フロー:
        1. CLI 引数パース（--screenshots-dir, --output）
        2. PNG ファイルをファイル名昇順で収集
        3. Workbook 作成（デフォルトシートを削除）
        4. 各 PNG に対してシート生成
        5. 出力パスの親ディレクトリを mkdir(exist_ok=True) で保証
        6. Workbook を保存
        7. 保存先パスと処理件数をコンソール出力
    """
    args = parse_args()
    screenshots_dir = Path(args.screenshots_dir)
    output_path = Path(args.output)

    # PNG ファイル一覧を収集する（エラー時は sys.exit で終了）
    png_files = collect_png_files(screenshots_dir)

    # Workbook を生成し、自動作成されるデフォルトシート "Sheet" を削除する。
    # openpyxl は Workbook 作成時に必ず 1 枚シートを生成するため、明示的に除去する。
    wb = Workbook()
    wb.remove(wb.active)

    for png_path in png_files:
        add_screenshot_sheet(wb, png_path, SHEET_NAME_MAPPING)

    # 出力先ディレクトリが存在しない場合は作成する
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)

    print(f"保存完了: {output_path}（{len(png_files)} シート）")


if __name__ == "__main__":
    main()
